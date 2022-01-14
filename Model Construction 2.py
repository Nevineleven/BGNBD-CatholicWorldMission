#!/usr/bin/env python
# coding: utf-8

# In[138]:


import csv
import pandas as pd
import datetime as dt
from datetime import datetime
from itertools import islice
import numpy as np
from openpyxl import load_workbook
import collections
import openpyxl as op
import math
import xlsxwriter
from openpyxl import load_workbook


# In[139]:


exportFile = 'Model construction CWM bgnbd2.xlsx'
writer = pd.ExcelWriter(exportFile, engine = 'openpyxl')
grgFile = "GRG Running.xlsx"

excelBook = load_workbook(exportFile)
del excelBook["BGNBD Estimation"]
writer.book = excelBook
writer.sheets = dict((ws.title, ws) for ws in excelBook.worksheets)


# In[140]:


df_GRG = pd.read_excel(grgFile, header=None, nrows = 5, usecols = [0,1])
df_BGNBD = pd.read_excel(grgFile, skiprows=6)
df_BGNBD


# In[141]:


r = df_GRG.iloc[0,1]
alpha = df_GRG.iloc[1,1]
a = df_GRG.iloc[2,1]
b = df_GRG.iloc[3,1]


# In[142]:


df_input = pd.read_csv("Input.csv")
inputList = df_input["Values"].tolist()
df_input

beginning_date = dt.datetime(int (inputList[1]), int (inputList[2]), int (inputList[3]))
predict_thru = dt.datetime(int (inputList[4]), int (inputList[5]), int (inputList[6]))

def datevalue(date1):
    temp = dt.datetime(1899, 12, 30)    # Note, not 31st Dec but 30th!
    delta = date1 - temp
    return float(delta.days) + (float(delta.seconds) / 86400)

def valuedate(fh):
    fh = int(fh)
    return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + fh - 2)

beginning_date = datevalue(beginning_date)
predict_thru = datevalue(predict_thru)

endDate = beginning_date + .5*(predict_thru-beginning_date)
firstTimeTo = beginning_date + .31 * .5 * (predict_thru-beginning_date)
print(endDate)


# In[143]:


df_ext = pd.DataFrame()
tuntil = (predict_thru - beginning_date + 1)/365
print(tuntil)
n = 1/365
t = [n]
while n <= tuntil:
    n = n + (1/365)
    t.append(n)
df_ext['t'] = t
df_ext['E(X(t))'] = 0
df_ext['2F1'] = 0
df_ext


# In[144]:


z=[]
for index, rows in islice(df_ext.iterrows(), 0, None):
    z.append(df_ext.iloc[index,0] / (alpha + df_ext.iloc[index,0]))
df_ext['z']=z
df_ext['0'] = 1
df_ext


# In[145]:


count = 1
ext_cols = []
extCond = True
while (extCond):
    for index, rows in islice(df_ext.iterrows(), 0, None):
        ext_cols.append(df_ext.iloc[index, count+3]*df_ext.iloc[index, 3]*(r+count-1)*(b+count-1)/((a+b-1+count-1)*count))
    if (ext_cols[-1] < 9e-11):
        extCond = False
    else:
        df_ext[str(count)] = ext_cols
        count = count + 1
        ext_cols = [];
df_ext


# In[146]:


df_save = df_ext
df_save


# In[147]:


df_ext = df_save
ext = []
df_ext['2F1'] = df_ext.iloc[:,4:-1].sum(axis=1)
extConst = (a+b-1)/(a-1)
for index, rows in islice(df_ext.iterrows(), 0, None):
    ext.append(extConst*(1-pow((alpha/(alpha+df_ext.iloc[index,0])),r) * df_ext.iloc[index,2]))
df_ext['E(X(t))'] = ext
df_ext


# In[148]:


calibLength = (endDate-beginning_date+1)/365
df_ns = pd.DataFrame()
df_ns["Constituent ID"] = df_BGNBD["Constituent ID"]
df_ns["T (total time span)"] = df_BGNBD["T (total time span)"]


# In[149]:


firstPurch = []
T = df_ns['T (total time span)'].to_list()
for index, row in islice(df_ns.iterrows(), 0, None):
    firstPurch.append(calibLength - T[index])
df_ns['time of 1st purchase'] = (firstPurch)
maxTime = df_ns['time of 1st purchase'].max()
df_ns


# In[150]:


weekUntil = (predict_thru - beginning_date)/7
df_sls_top = pd.DataFrame()
trialDon = []
numDon = []
i = 1/365
while (i<=maxTime):
    trialDon.append(i)
    numDon.append(df_ns.loc[(df_ns['time of 1st purchase'] >= i-.001) & (df_ns['time of 1st purchase'] < i+.001)].count()[0])
    i = i + 1/365
df_sls_top["Time of Trial Donation"] = trialDon
df_sls_top["Number of Donors"] = numDon
df_sls_top


# In[151]:


df_sls = pd.DataFrame()
df_sls['t'] = t
df_sls['Cum. Rpt.'] = ""
df_sls['E(X(t))'] = ext
df_sls


# In[152]:


slsCount = 0
slsCols = []
while (slsCount < df_sls_top.count()[0]):
    for index, rows in islice(df_sls.iterrows(), 0, None):
        if (df_sls.iloc[index, 0] <= df_sls_top["Time of Trial Donation"][slsCount]):
            slsCols.append(0)
        else:
            indx = 365*(df_sls.iloc[index, 0]-df_sls_top["Time of Trial Donation"][slsCount])
            slsCols.append(ext[round(indx)-1])
    df_sls[str(slsCount)] = slsCols
    slsCount = slsCount+1
    slsCols = []
df_sls


# In[153]:


df_save1 = df_sls


# In[154]:


# cumRpt = []
# cumColCnt = 0
# #for c in range(df_sls_top.count()[0]):
# print(df_sls.iloc[0, 3:-1])
# print(df_sls_top["Time of Trial Donation"])
# print(df_sls.iloc[0, 3:-1] * df_sls_top["Time of Trial Donation"])
# #for r in range (df_sls.count()[0]):
#     #print((df_sls.iloc[r, 3:-1] * df_sls_top["Time of Trial Donation"]))
#     #cumRpt.append(sum(df_sls.iloc[r, 3:-1] * df_sls_top["Time of Trial Donation"]))
# df_sls["Cum. Rpt"] = cumRpt
# df_sls
cumRpt = []
cumRptSum = 0
#print(df_sls_top["Number of Donors"])
#print(df_sls.iloc[1,3:-1].astype('float64'))
#print (df_sls_top["Number of Donors"].multiply((df_sls.iloc[0,3:-1].astype('float64')), level = 0))
for index, rows in islice(df_sls.iterrows(), 0, None):
    i = 0
    cumRptSum = 0
    #print(cumRptSum + df_sls.iloc[index, i+3])
    #print(numDon[i])
    while (i<df_sls_top.count()[0]): 
        cumRptSum = cumRptSum + df_sls.iloc[index, i+3] * numDon[i]
        i = i+1
    cumRpt.append(cumRptSum)
    #print(cumRptSum)
df_sls["Cum. Rpt."] = cumRpt
df_sls


# In[155]:


df_checkCumRpt = pd.DataFrame()
week = list(range(1, int(weekUntil)+1))
df_checkCumRpt["week"] = week
df_checkCumRpt


# In[156]:


checkDon = [cumRpt[6]]
checkWeekDon = [checkDon[0]]
weekCount = 2
while (weekCount <= weekUntil):
    checkDon.append(cumRpt[7*weekCount-1])
    checkWeekDon.append(checkDon[weekCount-1] - checkDon[weekCount-2])
    weekCount = weekCount + 1
df_checkCumRpt["Cum. Rpt. Dons."] = checkDon
df_checkCumRpt["Weekly Rpt. Dons."] = checkWeekDon
df_checkCumRpt


# In[157]:


startDate = [beginning_date]
endDate = [startDate[0] + 6]
for index, row in islice(df_checkCumRpt.iterrows(), 1, None):
    startDate.append(startDate[index-1]+7)
    endDate.append(endDate[index-1]+7)
df_checkCumRpt['week start'] = (startDate)
df_checkCumRpt['end'] = (endDate)
df_checkCumRpt


# In[158]:


t = [0]
t2 = [6/365]
for index, row in islice(df_checkCumRpt.iterrows(), 1, None):
    t.append(t[index-1]+7/365)
    t2.append(t2[index-1]+7/365)
df_checkCumRpt["num rpt don"] = ""
df_checkCumRpt["cumul."] = ""
df_checkCumRpt['t'] = (t)
df_checkCumRpt['t2'] = (t2)
df_checkCumRpt


# In[159]:


dfraw2 = pd.read_excel(exportFile, sheet_name = "Raw data2")
dfraw2


# In[160]:


numRpt = []
cummul = []
weekStart = df_checkCumRpt['week start'].tolist()
weekEnd = df_checkCumRpt['end'].tolist()
t = df_checkCumRpt['t'].tolist()
t2 = df_checkCumRpt['t2'].tolist()
count99 = 0
countTotal99 = 0
for index, row in islice(df_checkCumRpt.iterrows(), 0, None):
    count99 = count99 + len(dfraw2.loc[(dfraw2['Gift Date']>=weekStart[index]) & (dfraw2['Gift Date']<=weekEnd[index])])
    count99 = count99 - len(df_ns.loc[df_ns['time of 1st purchase']>=(t[index]-.001)])
    count99 = count99 + len(df_ns.loc[df_ns['time of 1st purchase']>(t2[index]+.001)])
    countTotal99 = countTotal99+count99
    numRpt.append(count99)
    cummul.append(countTotal99)
    count99 = 0
df_checkCumRpt["num rpt don"] = numRpt
df_checkCumRpt["cumul."] = cummul
df_checkCumRpt


# In[161]:


df_allCondExp = df_BGNBD.iloc[:,0:5].copy(deep=True)
endDate = beginning_date + .5*(predict_thru-beginning_date)
constT = (predict_thru - endDate)/365
df_allCondExp['t'] = constT
df_allCondExp["E(Y(t)|X=x,t_x,T)"] = ""
df_allCondExp["2F1"] = 0
df_allCondExp["a"] = df_allCondExp.iloc[:,2]+r
df_allCondExp["b"] = df_allCondExp.iloc[:,2] + b
df_allCondExp["c"] = df_allCondExp.iloc[:,2] + b + a - 1
df_allCondExp["z"] = constT/(alpha + df_allCondExp.iloc[:,4] + constT)
df_allCondExp['0'] = 1
df_allCondExp


# In[162]:


aceCount = 1
aceArr = []
while (aceCount <= 200):
    df_allCondExp[str(aceCount)]=df_allCondExp.iloc[:,aceCount+11]*(df_allCondExp.iloc[:,8]+aceCount-1)*(df_allCondExp.iloc[:,9]+aceCount-1)/((df_allCondExp.iloc[:,10]+aceCount-1)*aceCount)*df_allCondExp.iloc[:,11]
    aceCount = aceCount + 1
df_allCondExp


# In[163]:


aceCount1 = 0
while (aceCount1 <= 200):
    df_allCondExp["2F1"] = df_allCondExp["2F1"]+df_allCondExp.iloc[:,aceCount1+12]
    aceCount1 = aceCount1 + 1
df_allCondExp


# In[164]:


longNameArr = []
#((a+b+df_allCondExp.iloc[:,2]-1)/(a-1)*(1-pow(((alpha+df_allCondExp.iloc[:,4])/(alpha+df_allCondExp.iloc[:,4]+df_allCondExp.iloc[:,5])),(r+df_allCondExp.iloc[:,2]))*df_allCondExp.iloc[:,7])/(1+(df_allCondExp.iloc[:,2]>0)*a/(b+df_allCondExp.iloc[:,2]-1)*pow(((alpha+df_allCondExp.iloc[:,4])/(alpha+df_allCondExp.iloc[:,3])),(r+df_allCondExp.iloc[:,2]))))
for index, row in islice(df_allCondExp.iterrows(), 0, None):
    longNameArr.append((a+b+df_allCondExp.iloc[index,2]-1)/(a-1)*(1-pow(((alpha+df_allCondExp.iloc[index,4])/(alpha+df_allCondExp.iloc[index,4]+df_allCondExp.iloc[index,5])),(r+df_allCondExp.iloc[index,2]))*df_allCondExp.iloc[index,7])/(1+(df_allCondExp.iloc[index,2]>0)*a/(b+df_allCondExp.iloc[index,2]-1)*pow(((alpha+df_allCondExp.iloc[index,4])/(alpha+df_allCondExp.iloc[index,3])),(r+df_allCondExp.iloc[index,2]))))
df_allCondExp["E(Y(t)|X=x,t_x,T)"] = longNameArr
df_allCondExp


# In[165]:


df_pAlive = df_BGNBD.iloc[:,0:5].copy(deep=True)
paliveinfo = []
for index, row in islice(df_allCondExp.iterrows(), 0, None):
    paliveinfo.append(1/(1+(df_pAlive.iloc[index,2]>0)*(a/(b+df_pAlive.iloc[index,2]-1))*(pow(((alpha+df_pAlive.iloc[index,4])/(alpha+df_pAlive.iloc[index,3])),(r+df_pAlive.iloc[index,2])))))
df_pAlive["P(Alive) Info"] = paliveinfo
df_pAlive


# In[166]:


dfdp2 = pd.read_excel(exportFile, sheet_name = "Data Prep2")
dfdp2


# In[167]:


actNonlap = []
df_pAlive["Active"] = ~dfdp2["Inactive?"]*1
for index, row in islice(df_pAlive.iterrows(), 0, None):
    if df_pAlive.iloc[index,6] == 0:
        actNonlap.append(0)
    else:
        if (df_pAlive.iloc[index,4]-df_pAlive.iloc[index,3])>2:
            actNonlap.append(0)
        else:
            actNonlap.append(1)
df_pAlive["Act/Nonlap"] = actNonlap
df_pAlive


# In[168]:


df_onlyx = df_pAlive.copy(deep=True)
df_onlyx


# In[169]:


dfGRG.to_excel(writer, "BGNBD Estimation", index = False, header = False, startrow=0, startcol=0)
df_BGNBD.to_excel(writer, "BGNBD Estimation", index = False, startrow = 6, startcol=0)
df_ext.to_excel(writer, "E(X(t))", index = False)
df_ns.to_excel(writer, "n_s", index=False)
df_sls.to_excel(writer, "CumRptSls", index=False)
df_checkCumRpt.to_excel(writer, "Check CumRpt", index=False)
df_allCondExp.to_excel(writer, "All Cond. Exp.", index=False)
df_pAlive.to_excel(writer, "P(alive)", index=False)
df_onlyx.to_excel(writer, "only x>0", index=False)
writer.save()


# In[170]:


f= open("ModelCon.txt","w+")
f.write("Model Construction Successfully Completed")
f.close()

